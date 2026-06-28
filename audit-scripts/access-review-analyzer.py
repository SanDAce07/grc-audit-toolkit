"""
Access Review Analyzer
======================
IT Audit Tool — GRC Toolkit
Author: Sandesh | ULM Accounting & CIS

Purpose:
    Analyzes a user access export CSV and flags access control anomalies.
    Outputs a formatted Excel workbook with findings and summary.

Input CSV columns (required):
    user_id, name, email, department, role, system, access_level,
    last_login, status, mfa_enabled

Usage:
    python access-review-analyzer.py --input user_access.csv --output access_review_findings.xlsx
    python access-review-analyzer.py  (uses sample data if no file provided)
"""

import pandas as pd
import argparse
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

REQUIRED_COLUMNS = [
    "user_id", "name", "email", "department", "role",
    "system", "access_level", "last_login", "status", "mfa_enabled"
]

def validate_input(df):
    """Check required columns exist and warn about data quality issues."""
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"Input CSV is missing required column(s): {', '.join(missing)}\n"
            f"Required columns: {', '.join(REQUIRED_COLUMNS)}"
        )
    # Warn about nulls in key columns
    for col in ["user_id", "name", "status", "system"]:
        null_count = df[col].isna().sum()
        if null_count > 0:
            print(f"   ⚠️  Warning: {null_count} null value(s) in column '{col}' — these rows may be skipped.")
    return True


# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
DORMANT_DAYS = 90          # Flag accounts with no login in this many days
PRIVILEGED_ROLES = [       # Roles considered high-privilege
    "admin", "administrator", "superuser", "root",
    "privileged", "sysadmin", "dba", "owner"
]
HIGH_RISK_SYSTEMS = [      # Systems where findings are automatically elevated
    "ERP", "Financial System", "Payroll", "HR System", "Active Directory"
]


# ─────────────────────────────────────────────
# SAMPLE DATA GENERATOR
# ─────────────────────────────────────────────
def generate_sample_data():
    """Generate realistic sample user access data for demo/testing."""
    today = datetime.today()
    data = [
        # Normal active users
        {"user_id": "U001", "name": "Alice Johnson",    "email": "alice@company.com",   "department": "Finance",    "role": "Analyst",       "system": "ERP",              "access_level": "Read",       "last_login": today - timedelta(days=2),   "status": "Active",     "mfa_enabled": "Yes"},
        {"user_id": "U002", "name": "Bob Smith",        "email": "bob@company.com",     "department": "IT",         "role": "Administrator", "system": "Active Directory", "access_level": "Admin",      "last_login": today - timedelta(days=1),   "status": "Active",     "mfa_enabled": "Yes"},
        {"user_id": "U003", "name": "Carol White",      "email": "carol@company.com",   "department": "Accounting", "role": "Senior",        "system": "Financial System", "access_level": "Read/Write", "last_login": today - timedelta(days=5),   "status": "Active",     "mfa_enabled": "Yes"},
        # FINDING: Terminated user still active
        {"user_id": "U004", "name": "David Lee",        "email": "david@company.com",   "department": "Sales",      "role": "Manager",       "system": "ERP",              "access_level": "Read/Write", "last_login": today - timedelta(days=45),  "status": "Terminated", "mfa_enabled": "Yes"},
        # FINDING: Dormant account (no login > 90 days)
        {"user_id": "U005", "name": "Eva Martinez",     "email": "eva@company.com",     "department": "HR",         "role": "Analyst",       "system": "HR System",        "access_level": "Read",       "last_login": today - timedelta(days=120), "status": "Active",     "mfa_enabled": "Yes"},
        # FINDING: Privileged user without MFA
        {"user_id": "U006", "name": "Frank Brown",      "email": "frank@company.com",   "department": "IT",         "role": "Sysadmin",      "system": "Active Directory", "access_level": "Admin",      "last_login": today - timedelta(days=3),   "status": "Active",     "mfa_enabled": "No"},
        # FINDING: Excessive access (non-IT user with admin)
        {"user_id": "U007", "name": "Grace Kim",        "email": "grace@company.com",   "department": "Finance",    "role": "Analyst",       "system": "ERP",              "access_level": "Admin",      "last_login": today - timedelta(days=10),  "status": "Active",     "mfa_enabled": "Yes"},
        # Normal users
        {"user_id": "U008", "name": "Henry Wilson",     "email": "henry@company.com",   "department": "IT",         "role": "Developer",     "system": "ERP",              "access_level": "Read",       "last_login": today - timedelta(days=7),   "status": "Active",     "mfa_enabled": "Yes"},
        # FINDING: Terminated + dormant
        {"user_id": "U009", "name": "Iris Chen",        "email": "iris@company.com",    "department": "Finance",    "role": "Manager",       "system": "Payroll",          "access_level": "Read/Write", "last_login": today - timedelta(days=200), "status": "Terminated", "mfa_enabled": "No"},
        # FINDING: No MFA on financial system
        {"user_id": "U010", "name": "Jack Taylor",      "email": "jack@company.com",    "department": "Accounting", "role": "Analyst",       "system": "Financial System", "access_level": "Read/Write", "last_login": today - timedelta(days=4),   "status": "Active",     "mfa_enabled": "No"},
        # Normal
        {"user_id": "U011", "name": "Karen Davis",      "email": "karen@company.com",   "department": "IT",         "role": "Analyst",       "system": "Active Directory", "access_level": "Read",       "last_login": today - timedelta(days=1),   "status": "Active",     "mfa_enabled": "Yes"},
        # FINDING: Duplicate access (same user, same system, different records)
        {"user_id": "U012", "name": "Leo Garcia",       "email": "leo@company.com",     "department": "Finance",    "role": "Senior",        "system": "ERP",              "access_level": "Read",       "last_login": today - timedelta(days=3),   "status": "Active",     "mfa_enabled": "Yes"},
        {"user_id": "U012", "name": "Leo Garcia",       "email": "leo@company.com",     "department": "Finance",    "role": "Admin",         "system": "ERP",              "access_level": "Admin",      "last_login": today - timedelta(days=3),   "status": "Active",     "mfa_enabled": "Yes"},
        # Normal
        {"user_id": "U013", "name": "Mia Thompson",     "email": "mia@company.com",     "department": "Accounting", "role": "Junior",        "system": "Financial System", "access_level": "Read",       "last_login": today - timedelta(days=2),   "status": "Active",     "mfa_enabled": "Yes"},
        {"user_id": "U014", "name": "Nathan Scott",     "email": "nathan@company.com",  "department": "IT",         "role": "DBA",           "system": "ERP",              "access_level": "Admin",      "last_login": today - timedelta(days=1),   "status": "Active",     "mfa_enabled": "Yes"},
        # FINDING: Contractor with excessive access
        {"user_id": "U015", "name": "Olivia Nguyen",    "email": "olivia@vendor.com",   "department": "Contractor", "role": "Contractor",    "system": "Payroll",          "access_level": "Read/Write", "last_login": today - timedelta(days=15),  "status": "Active",     "mfa_enabled": "No"},
    ]
    return pd.DataFrame(data)


# ─────────────────────────────────────────────
# ANOMALY DETECTION ENGINE
# ─────────────────────────────────────────────
def analyze_access(df):
    """Run all anomaly checks and return a findings dataframe."""
    findings = []
    today = datetime.today()

    # Resilient last_login parsing — coerce bad values to NaT
    df["last_login"] = pd.to_datetime(df["last_login"], errors="coerce")

    # For NaT last_login, treat as maximally dormant (9999 days)
    df["days_since_login"] = (today - df["last_login"]).dt.days.fillna(9999).astype(int)
    df["days_since_login"] = df["days_since_login"].clip(lower=0)  # no negative values

    for _, row in df.iterrows():
        user_findings = []

        # ── Check 1: Terminated user still has access ──
        if str(row["status"]).strip().lower() == "terminated":
            user_findings.append({
                "finding_type": "Terminated User — Active Access",
                "description": f"{row['name']} is terminated but retains {row['access_level']} access to {row['system']}.",
                "risk": "Critical",
                "recommendation": "Immediately revoke all system access. Verify offboarding procedure was followed."
            })

        # ── Check 2: Dormant account ──
        elif row["days_since_login"] > DORMANT_DAYS:
            risk = "High" if row["system"] in HIGH_RISK_SYSTEMS else "Medium"
            user_findings.append({
                "finding_type": "Dormant Account",
                "description": f"{row['name']} has not logged into {row['system']} in {row['days_since_login']} days (threshold: {DORMANT_DAYS} days).",
                "risk": risk,
                "recommendation": "Disable account or require re-verification. Confirm user still requires access."
            })

        # ── Check 3: MFA not enabled ──
        if str(row["mfa_enabled"]).strip().lower() in ["no", "false", "0", ""]:
            is_privileged = any(p in str(row["role"]).lower() for p in PRIVILEGED_ROLES)
            is_high_risk_system = row["system"] in HIGH_RISK_SYSTEMS
            if is_privileged or is_high_risk_system:
                risk = "Critical" if is_privileged else "High"
                user_findings.append({
                    "finding_type": "MFA Not Enabled — Privileged/Sensitive System",
                    "description": f"{row['name']} ({row['role']}) does not have MFA enabled on {row['system']}.",
                    "risk": risk,
                    "recommendation": "Enforce MFA immediately. Update access control policy to require MFA for all privileged and sensitive system access."
                })
            else:
                user_findings.append({
                    "finding_type": "MFA Not Enabled",
                    "description": f"{row['name']} does not have MFA enabled on {row['system']}.",
                    "risk": "Medium",
                    "recommendation": "Enable MFA as part of standard access provisioning."
                })

        # ── Check 4: Excessive privilege (non-IT with Admin) ──
        dept = str(row["department"]).strip().lower()
        access = str(row["access_level"]).strip().lower()
        role = str(row["role"]).strip().lower()
        is_privileged_role = any(p in role for p in PRIVILEGED_ROLES)
        is_it_dept = dept in ["it", "information technology", "cybersecurity", "security"]

        if "admin" in access and not is_it_dept and not is_privileged_role:
            user_findings.append({
                "finding_type": "Excessive Privilege — Non-IT Admin Access",
                "description": f"{row['name']} ({row['department']}) has Admin-level access to {row['system']}, which may violate least privilege principles.",
                "risk": "High",
                "recommendation": "Review and remediate. Reduce access to minimum required. Obtain management approval if Admin access is justified."
            })

        # ── Check 5: Contractor with elevated access ──
        is_contractor = dept == "contractor"
        has_elevated  = ("read/write" in access) or ("admin" in access)
        if is_contractor and has_elevated:
            user_findings.append({
                "finding_type": "Contractor — Elevated Access",
                "description": f"Contractor {row['name']} has {row['access_level']} access to {row['system']}. Contractor access should be limited and time-bound.",
                "risk": "High",
                "recommendation": "Verify business justification. Ensure access is time-limited and logged. Apply least privilege."
            })

        # Compile findings for this user
        for f in user_findings:
            findings.append({
                "User ID":        row["user_id"],
                "Name":           row["name"],
                "Email":          row["email"],
                "Department":     row["department"],
                "Role":           row["role"],
                "System":         row["system"],
                "Access Level":   row["access_level"],
                "Last Login":     row["last_login"].strftime("%Y-%m-%d"),
                "Days Since Login": int(row["days_since_login"]),
                "Status":         row["status"],
                "MFA Enabled":    row["mfa_enabled"],
                "Finding Type":   f["finding_type"],
                "Risk Rating":    f["risk"],
                "Description":    f["description"],
                "Recommendation": f["recommendation"],
            })

    # ── Check 6: Duplicate access (same user_id + system) ──
    dupes = df[df.duplicated(subset=["user_id", "system"], keep=False)]
    seen_dupes = set()
    for _, row in dupes.iterrows():
        key = (row["user_id"], row["system"])
        if key not in seen_dupes:
            seen_dupes.add(key)
            findings.append({
                "User ID":        row["user_id"],
                "Name":           row["name"],
                "Email":          row["email"],
                "Department":     row["department"],
                "Role":           row["role"],
                "System":         row["system"],
                "Access Level":   "Multiple",
                "Last Login":     row["last_login"].strftime("%Y-%m-%d"),
                "Days Since Login": int(row["days_since_login"]),
                "Status":         row["status"],
                "MFA Enabled":    row["mfa_enabled"],
                "Finding Type":   "Duplicate Access — Same User, Same System",
                "Risk Rating":    "High",
                "Description":    f"{row['name']} (ID: {row['user_id']}) has multiple access records for {row['system']}. This may indicate conflicting roles or orphaned entries.",
                "Recommendation": "Review all access records for this user. Remove duplicate or conflicting entitlements. Ensure single, appropriate role assignment.",
            })

    return pd.DataFrame(findings) if findings else pd.DataFrame()


# ─────────────────────────────────────────────
# EXCEL REPORT WRITER
# ─────────────────────────────────────────────
RISK_COLORS = {
    "Critical": "C00000",  # Dark red
    "High":     "FF0000",  # Red
    "Medium":   "FFC000",  # Amber
    "Low":      "92D050",  # Green
}

def style_workbook(output_path, df_all, df_findings):
    """Apply professional formatting to the Excel workbook."""
    wb = load_workbook(output_path)

    header_fill   = PatternFill("solid", fgColor="1F3864")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def format_sheet(ws, col_widths):
        # Header row
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border
        # Data rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border    = thin_border
                cell.alignment = left_align
        # Column widths
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        ws.row_dimensions[1].height = 30

    # ── Sheet 1: All Users ──
    ws1 = wb["All Users"]
    format_sheet(ws1, [10, 20, 28, 15, 18, 22, 15, 14, 16, 12, 12])
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ── Sheet 2: Findings ──
    ws2 = wb["Findings"]
    format_sheet(ws2, [10, 20, 28, 15, 18, 22, 14, 14, 16, 12, 12, 32, 12, 45, 50])
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    # Color-code Risk Rating column (col 13 = M)
    risk_col = None
    for i, cell in enumerate(ws2[1], 1):
        if cell.value == "Risk Rating":
            risk_col = i
            break
    if risk_col:
        for row in ws2.iter_rows(min_row=2):
            risk_cell = row[risk_col - 1]
            color = RISK_COLORS.get(str(risk_cell.value), None)
            if color:
                risk_cell.fill = PatternFill("solid", fgColor=color)
                risk_cell.font = Font(bold=True, color="FFFFFF" if risk_cell.value in ["Critical", "High"] else "000000")
                risk_cell.alignment = center_align

    # ── Sheet 3: Summary ──
    ws3 = wb["Summary"]
    format_sheet(ws3, [35, 12, 12, 12, 12, 12])
    ws3.freeze_panes = "A2"

    # Color summary risk counts
    for row in ws3.iter_rows(min_row=2):
        for cell in row:
            if cell.column > 1 and cell.value and isinstance(cell.value, int) and cell.value > 0:
                header_val = ws3.cell(1, cell.column).value
                color = RISK_COLORS.get(str(header_val), None)
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(bold=True, color="FFFFFF" if header_val in ["Critical", "High"] else "000000")

    wb.save(output_path)


def write_report(df_all, df_findings, output_path):
    """Write findings to a formatted Excel workbook."""

    # Summary table
    if not df_findings.empty:
        summary_data = (
            df_findings.groupby(["Finding Type", "Risk Rating"])
            .size()
            .reset_index(name="Count")
            .pivot(index="Finding Type", columns="Risk Rating", values="Count")
            .fillna(0)
            .astype(int)
            .reset_index()
        )
        # Ensure all risk columns exist
        for col in ["Critical", "High", "Medium", "Low"]:
            if col not in summary_data.columns:
                summary_data[col] = 0
        summary_data["Total"] = summary_data[["Critical", "High", "Medium", "Low"]].sum(axis=1)
        summary_data = summary_data[["Finding Type", "Critical", "High", "Medium", "Low", "Total"]]
    else:
        summary_data = pd.DataFrame(columns=["Finding Type", "Critical", "High", "Medium", "Low", "Total"])

    # Prepare All Users sheet (drop calculated column)
    df_display = df_all.copy()
    df_display["last_login"] = pd.to_datetime(df_display["last_login"]).dt.strftime("%Y-%m-%d")
    df_display = df_display.drop(columns=["days_since_login"], errors="ignore")
    df_display.columns = [c.replace("_", " ").title() for c in df_display.columns]

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_display.to_excel(writer, sheet_name="All Users", index=False)
        if not df_findings.empty:
            df_findings.to_excel(writer, sheet_name="Findings", index=False)
        else:
            # Write empty sheet with correct headers so it looks intentional
            empty_findings = pd.DataFrame(columns=[
                "User ID", "Name", "Email", "Department", "Role", "System",
                "Access Level", "Last Login", "Days Since Login", "Status",
                "MFA Enabled", "Finding Type", "Risk Rating", "Description", "Recommendation"
            ])
            empty_findings.to_excel(writer, sheet_name="Findings", index=False)
        summary_data.to_excel(writer, sheet_name="Summary", index=False)

    style_workbook(output_path, df_all, df_findings)
    return summary_data


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    global DORMANT_DAYS, PRIVILEGED_ROLES, HIGH_RISK_SYSTEMS
    parser = argparse.ArgumentParser(description="Access Review Analyzer — IT Audit Tool")
    parser.add_argument("--input",  "-i", help="Path to user access CSV file", default=None)
    parser.add_argument("--output", "-o", help="Output Excel file path",       default="access_review_findings.xlsx")
    parser.add_argument("--dormant-days",      type=int, default=DORMANT_DAYS,
                        help=f"Days without login to flag as dormant (default: {DORMANT_DAYS})")
    parser.add_argument("--privileged-roles",  default=",".join(PRIVILEGED_ROLES),
                        help="Comma-separated privileged role keywords (overrides built-in list)")
    parser.add_argument("--high-risk-systems", default=",".join(HIGH_RISK_SYSTEMS),
                        help="Comma-separated high-risk system names (overrides built-in list)")
    args = parser.parse_args()

    # Apply CLI overrides to globals
    DORMANT_DAYS      = args.dormant_days
    PRIVILEGED_ROLES  = [r.strip().lower() for r in args.privileged_roles.split(",")]
    HIGH_RISK_SYSTEMS = [s.strip() for s in args.high_risk_systems.split(",")]

    print("=" * 60)
    print("  ACCESS REVIEW ANALYZER — GRC & IT Audit Toolkit")
    print("=" * 60)

    # Load data
    if args.input:
        print(f"\n📂 Loading: {args.input}")
        df = pd.read_csv(args.input)
        try:
            validate_input(df)
        except ValueError as e:
            print(f"\n❌ ERROR: {e}")
            return
    else:
        print("\n⚠️  No input file provided. Using sample data for demo.")
        df = generate_sample_data()

    print(f"   {len(df)} user access records loaded.")

    # Run analysis
    print("\n🔍 Running anomaly checks...")
    print(f"   ✓ Terminated users with active access")
    print(f"   ✓ Dormant accounts (>{DORMANT_DAYS} days inactive)")
    print(f"   ✓ MFA not enabled (privileged/sensitive systems)")
    print(f"   ✓ Excessive privilege (non-IT admin access)")
    print(f"   ✓ Contractor elevated access")
    print(f"   ✓ Duplicate access records")

    df_findings = analyze_access(df)

    # Write report
    print(f"\n📊 Writing report: {args.output}")
    summary = write_report(df, df_findings, args.output)

    # Print console summary
    print("\n" + "=" * 60)
    print("  FINDINGS SUMMARY")
    print("=" * 60)
    if not df_findings.empty:
        total = len(df_findings)
        for risk in ["Critical", "High", "Medium", "Low"]:
            count = len(df_findings[df_findings["Risk Rating"] == risk])
            if count > 0:
                print(f"   {risk:<12}: {count}")
        print(f"   {'TOTAL':<12}: {total}")
        print(f"\n   Affected Users : {df_findings['User ID'].nunique()}")
        print(f"   Systems Flagged: {df_findings['System'].nunique()}")
    else:
        print("   ✅ No anomalies detected.")

    print(f"\n✅ Report saved: {args.output}")
    print("   Sheets: All Users | Findings | Summary")
    print("=" * 60)


if __name__ == "__main__":
    main()
