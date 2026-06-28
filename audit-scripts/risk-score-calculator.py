"""
Risk Score Calculator
=====================
IT Audit Tool — GRC Toolkit
Author: Sandesh | ULM Accounting & CIS

Purpose:
    Ingests a risk register CSV, calculates inherent and residual risk scores,
    ranks risks by priority, flags exceptions, and outputs a formatted Excel
    workbook with a heat map and executive summary.

Scoring Model:
    Likelihood  : 1 (Rare) → 5 (Almost Certain)
    Impact      : 1 (Negligible) → 5 (Critical)
    Inherent Score  = Likelihood × Impact  (max 25, before controls)
    Residual Score  = Inherent × (1 − Control Effectiveness)

    Control Effectiveness:
        none   → 0.00  (no controls)
        low    → 0.25  (weak/informal controls)
        medium → 0.50  (partially effective)
        high   → 0.75  (strong controls)
        full   → 1.00  (fully mitigated)

    Risk Rating:
        1–4   → Low
        5–9   → Medium
        10–19 → High
        20–25 → Critical

Input CSV columns (required):
    risk_id, risk_description, category, likelihood, impact,
    control_description, control_effectiveness, owner,
    mitigation_due_date, status

Usage:
    python risk-score-calculator.py --input risk_register.csv
    python risk-score-calculator.py  (uses sample data)
    python risk-score-calculator.py --format csv --output my_risks
"""

import pandas as pd
import argparse
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
REQUIRED_COLUMNS = [
    "risk_id", "risk_description", "category", "likelihood", "impact",
    "control_description", "control_effectiveness", "owner",
    "mitigation_due_date", "status"
]

CONTROL_EFFECTIVENESS_MAP = {
    "none":   0.00,
    "low":    0.25,
    "medium": 0.50,
    "high":   0.75,
    "full":   1.00,
}

RISK_RATING_BANDS = [
    (1,  4,  "Low",      "92D050"),   # Green
    (5,  9,  "Medium",   "FFC000"),   # Amber
    (10, 19, "High",     "FF0000"),   # Red
    (20, 25, "Critical", "C00000"),   # Dark Red
]

LIKELIHOOD_LABELS = {
    1: "Rare",
    2: "Unlikely",
    3: "Possible",
    4: "Likely",
    5: "Almost Certain",
}

IMPACT_LABELS = {
    1: "Negligible",
    2: "Minor",
    3: "Moderate",
    4: "Major",
    5: "Critical",
}


# ─────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────
def normalize_str(value):
    """Null-safe lowercase string normalization."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    return str(value).strip().lower()


def get_risk_rating(score):
    """Return (rating_label, hex_color) for a given score."""
    for low, high, label, color in RISK_RATING_BANDS:
        if low <= score <= high:
            return label, color
    return "Unknown", "FFFFFF"


def get_control_factor(effectiveness_str):
    """Convert control effectiveness string to reduction factor."""
    return CONTROL_EFFECTIVENESS_MAP.get(normalize_str(effectiveness_str), 0.0)


# ─────────────────────────────────────────────
# INPUT VALIDATION
# ─────────────────────────────────────────────
def validate_input(df):
    """Check required columns and warn about data issues."""
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"Input CSV is missing required column(s): {', '.join(missing)}\n"
            f"Required: {', '.join(REQUIRED_COLUMNS)}"
        )

    # Validate likelihood and impact are numeric 1-5
    for col in ["likelihood", "impact"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
        invalid = df[df[col].isna() | ~df[col].between(1, 5)]
        if not invalid.empty:
            print(f"   ⚠️  Warning: {len(invalid)} invalid value(s) in '{col}' "
                  f"(must be 1–5). These rows will be skipped in scoring.")

    # Validate control effectiveness
    valid_ce = set(CONTROL_EFFECTIVENESS_MAP.keys())
    invalid_ce = df[~df["control_effectiveness"].apply(normalize_str).isin(valid_ce)]
    if not invalid_ce.empty:
        print(f"   ⚠️  Warning: {len(invalid_ce)} invalid 'control_effectiveness' value(s). "
              f"Valid: {', '.join(valid_ce)}. Defaulting to 'none' (0%).")

    # Validate mitigation_due_date
    df["mitigation_due_date"] = pd.to_datetime(df["mitigation_due_date"], errors="coerce")
    invalid_dates = df[df["mitigation_due_date"].isna() & df["mitigation_due_date"].notna()]
    raw_invalid = df["mitigation_due_date"].isna().sum()
    if raw_invalid > 0:
        print(f"   ⚠️  Warning: {raw_invalid} missing/unparseable 'mitigation_due_date' value(s). "
              f"Overdue checks will be skipped for these rows.")
    return True


# ─────────────────────────────────────────────
# SAMPLE DATA GENERATOR
# ─────────────────────────────────────────────
def generate_sample_data():
    """Generate a realistic IT risk register with varied risk profiles."""
    today = datetime.today()
    data = [
        # Critical risks
        {
            "risk_id": "R-001", "risk_description": "Unauthorized access to financial systems via compromised credentials",
            "category": "Access Control", "likelihood": 4, "impact": 5,
            "control_description": "MFA enforced on all financial systems; quarterly access reviews",
            "control_effectiveness": "high", "owner": "CISO",
            "mitigation_due_date": (today + timedelta(days=30)).strftime("%Y-%m-%d"), "status": "Open"
        },
        {
            "risk_id": "R-002", "risk_description": "Ransomware attack encrypting critical business data",
            "category": "Cybersecurity", "likelihood": 3, "impact": 5,
            "control_description": "EDR solution deployed; backups maintained",
            "control_effectiveness": "medium", "owner": "IT Security",
            "mitigation_due_date": (today + timedelta(days=60)).strftime("%Y-%m-%d"), "status": "In Progress"
        },
        {
            "risk_id": "R-003", "risk_description": "ERP system outage during financial close period",
            "category": "Business Continuity", "likelihood": 2, "impact": 5,
            "control_description": "Disaster recovery plan; redundant systems",
            "control_effectiveness": "high", "owner": "IT Operations",
            "mitigation_due_date": (today + timedelta(days=90)).strftime("%Y-%m-%d"), "status": "Open"
        },
        # High risks
        {
            "risk_id": "R-004", "risk_description": "Unpatched vulnerabilities in internet-facing systems",
            "category": "Vulnerability Management", "likelihood": 4, "impact": 4,
            "control_description": "Monthly patching cycle; vulnerability scanning",
            "control_effectiveness": "medium", "owner": "IT Operations",
            "mitigation_due_date": (today - timedelta(days=15)).strftime("%Y-%m-%d"), "status": "Open"  # OVERDUE
        },
        {
            "risk_id": "R-005", "risk_description": "Insider threat — privileged user misuse of access",
            "category": "Access Control", "likelihood": 2, "impact": 5,
            "control_description": "PAM solution; activity logging",
            "control_effectiveness": "high", "owner": "CISO",
            "mitigation_due_date": (today + timedelta(days=45)).strftime("%Y-%m-%d"), "status": "Open"
        },
        {
            "risk_id": "R-006", "risk_description": "Third-party vendor breach exposing customer data",
            "category": "Third-Party Risk", "likelihood": 3, "impact": 4,
            "control_description": "Annual vendor risk assessments",
            "control_effectiveness": "low", "owner": "Procurement",
            "mitigation_due_date": (today - timedelta(days=30)).strftime("%Y-%m-%d"), "status": "Open"  # OVERDUE
        },
        {
            "risk_id": "R-007", "risk_description": "Phishing attacks leading to credential compromise",
            "category": "Cybersecurity", "likelihood": 5, "impact": 3,
            "control_description": "Security awareness training; email filtering",
            "control_effectiveness": "medium", "owner": "IT Security",
            "mitigation_due_date": (today + timedelta(days=20)).strftime("%Y-%m-%d"), "status": "In Progress"
        },
        {
            "risk_id": "R-008", "risk_description": "Change management failures causing production outages",
            "category": "ITGC", "likelihood": 3, "impact": 4,
            "control_description": "CAB process; change freeze periods",
            "control_effectiveness": "high", "owner": "Change Manager",
            "mitigation_due_date": (today + timedelta(days=75)).strftime("%Y-%m-%d"), "status": "Open"
        },
        # Medium risks
        {
            "risk_id": "R-009", "risk_description": "Inadequate backup testing leading to unrecoverable data loss",
            "category": "Business Continuity", "likelihood": 2, "impact": 4,
            "control_description": "Monthly backup jobs; annual restore test",
            "control_effectiveness": "medium", "owner": "IT Operations",
            "mitigation_due_date": (today + timedelta(days=120)).strftime("%Y-%m-%d"), "status": "Open"
        },
        {
            "risk_id": "R-010", "risk_description": "Non-compliance with data privacy regulations (GDPR/CCPA)",
            "category": "Compliance", "likelihood": 2, "impact": 4,
            "control_description": "Privacy policy; DPO appointed",
            "control_effectiveness": "medium", "owner": "Legal",
            "mitigation_due_date": (today + timedelta(days=180)).strftime("%Y-%m-%d"), "status": "Open"
        },
        {
            "risk_id": "R-011", "risk_description": "Shadow IT — unauthorized cloud services storing sensitive data",
            "category": "Data Governance", "likelihood": 4, "impact": 3,
            "control_description": "CASB solution partially deployed",
            "control_effectiveness": "low", "owner": "IT Security",
            "mitigation_due_date": (today + timedelta(days=60)).strftime("%Y-%m-%d"), "status": "In Progress"
        },
        # EXCEPTION: No owner assigned
        {
            "risk_id": "R-012", "risk_description": "Inadequate segregation of duties in financial approval workflow",
            "category": "ITGC", "likelihood": 3, "impact": 3,
            "control_description": "Workflow approval matrix under review",
            "control_effectiveness": "low", "owner": "",  # NO OWNER
            "mitigation_due_date": (today + timedelta(days=30)).strftime("%Y-%m-%d"), "status": "Open"
        },
        # EXCEPTION: No controls
        {
            "risk_id": "R-013", "risk_description": "Physical access to server room by unauthorized personnel",
            "category": "Physical Security", "likelihood": 2, "impact": 4,
            "control_description": "",  # NO CONTROLS
            "control_effectiveness": "none", "owner": "Facilities",
            "mitigation_due_date": (today - timedelta(days=5)).strftime("%Y-%m-%d"), "status": "Open"  # OVERDUE
        },
        # Low risks
        {
            "risk_id": "R-014", "risk_description": "Outdated software documentation increasing onboarding time",
            "category": "Operations", "likelihood": 4, "impact": 1,
            "control_description": "Documentation review scheduled quarterly",
            "control_effectiveness": "medium", "owner": "IT Manager",
            "mitigation_due_date": (today + timedelta(days=200)).strftime("%Y-%m-%d"), "status": "Open"
        },
        {
            "risk_id": "R-015", "risk_description": "Employee error in manual data entry causing report inaccuracies",
            "category": "Data Integrity", "likelihood": 3, "impact": 2,
            "control_description": "Dual-control review process; automated validation rules",
            "control_effectiveness": "high", "owner": "Finance Manager",
            "mitigation_due_date": (today + timedelta(days=150)).strftime("%Y-%m-%d"), "status": "Closed"
        },
    ]
    return pd.DataFrame(data)


# ─────────────────────────────────────────────
# SCORING ENGINE
# ─────────────────────────────────────────────
def calculate_scores(df):
    """Calculate inherent and residual risk scores for all risks."""
    df = df.copy()

    # Ensure numeric
    df["likelihood"] = pd.to_numeric(df["likelihood"], errors="coerce")
    df["impact"]     = pd.to_numeric(df["impact"],     errors="coerce")

    # Drop rows with invalid likelihood/impact
    valid = df["likelihood"].between(1, 5) & df["impact"].between(1, 5)
    skipped = (~valid).sum()
    if skipped > 0:
        print(f"   ⚠️  Skipping {skipped} row(s) with invalid likelihood/impact values.")
    df = df[valid].copy()

    # Inherent risk (before controls)
    df["inherent_score"] = df["likelihood"] * df["impact"]
    df["inherent_rating"] = df["inherent_score"].apply(lambda s: get_risk_rating(s)[0])

    # Control effectiveness
    df["control_factor"] = df["control_effectiveness"].apply(get_control_factor)

    # Residual risk (after controls)
    df["residual_score"]  = (df["inherent_score"] * (1 - df["control_factor"])).round(1)
    df["residual_rating"] = df["residual_score"].apply(lambda s: get_risk_rating(s)[0])

    # Risk reduction
    df["risk_reduction_%"] = ((df["control_factor"]) * 100).round(0).astype(int)

    # Readable labels
    df["likelihood_label"] = df["likelihood"].map(LIKELIHOOD_LABELS)
    df["impact_label"]     = df["impact"].map(IMPACT_LABELS)

    # Sort by residual score descending
    df = df.sort_values("residual_score", ascending=False).reset_index(drop=True)
    df["rank"] = range(1, len(df) + 1)

    return df


# ─────────────────────────────────────────────
# EXCEPTION DETECTION
# ─────────────────────────────────────────────
def detect_exceptions(df):
    """Flag risks with governance or control gaps."""
    today = datetime.today()
    exceptions = []

    for _, row in df.iterrows():
        rid = row["risk_id"]

        def flag(exc_type, severity, detail, rec):
            exceptions.append({
                "Risk ID":        rid,
                "Description":    row["risk_description"],
                "Category":       row["category"],
                "Residual Score": row["residual_score"],
                "Residual Rating":row["residual_rating"],
                "Exception Type": exc_type,
                "Severity":       severity,
                "Detail":         detail,
                "Recommendation": rec,
            })

        # 1. No owner assigned
        if normalize_str(row["owner"]) == "":
            flag(
                "No Owner Assigned", "High",
                f"Risk {rid} has no designated owner.",
                "Assign a risk owner immediately. Risk ownership is required for accountability and tracking."
            )

        # 2. No controls in place
        if normalize_str(row["control_description"]) == "" or \
                normalize_str(row["control_effectiveness"]) in ["", "none"]:
            flag(
                "No Controls in Place", "High",
                f"Risk {rid} has no compensating controls. Inherent score: {row['inherent_score']}.",
                "Design and implement controls. Prioritize based on inherent risk score."
            )

        # 3. Overdue mitigation
        if pd.notna(row["mitigation_due_date"]):
            due = pd.to_datetime(row["mitigation_due_date"])
            days_overdue = (today - due).days
            if days_overdue > 0 and normalize_str(row["status"]) not in ["closed", "resolved"]:
                severity = "Critical" if days_overdue > 30 else "High"
                flag(
                    "Overdue Mitigation", severity,
                    f"Risk {rid} mitigation was due {due.strftime('%Y-%m-%d')} "
                    f"({days_overdue} days ago). Status: {row['status']}.",
                    "Escalate to risk owner and management. Update mitigation plan with revised due date and justification."
                )

        # 4. High/Critical residual risk with weak controls
        if row["residual_score"] >= 10 and normalize_str(row["control_effectiveness"]) in ["none", "low"]:
            flag(
                "High Residual Risk — Weak Controls", "Critical",
                f"Risk {rid} has a residual score of {row['residual_score']} ({row['residual_rating']}) "
                f"with only {row['control_effectiveness']} control effectiveness.",
                "Immediately strengthen controls or accept risk with documented management approval. "
                "Consider additional compensating controls."
            )

        # 5. Open risk past its due date for more than 90 days
        if pd.notna(row["mitigation_due_date"]):
            due = pd.to_datetime(row["mitigation_due_date"])
            if (today - due).days > 90 and normalize_str(row["status"]) == "open":
                flag(
                    "Stale Open Risk — 90+ Days Overdue", "Critical",
                    f"Risk {rid} has been open and overdue for over 90 days.",
                    "Escalate to senior management. Risk must be mitigated, formally accepted, or transferred."
                )

    return pd.DataFrame(exceptions) if exceptions else pd.DataFrame()


# ─────────────────────────────────────────────
# HEAT MAP BUILDER
# ─────────────────────────────────────────────
def build_heat_map_data(df):
    """
    Build a 5x5 risk heat map matrix.
    Returns a dict mapping (likelihood, impact) -> list of risk_ids.
    """
    heat_map = {}
    for l in range(1, 6):
        for i in range(1, 6):
            heat_map[(l, i)] = []

    for _, row in df.iterrows():
        l = int(row["likelihood"])
        i = int(row["impact"])
        heat_map[(l, i)].append(row["risk_id"])

    return heat_map


# ─────────────────────────────────────────────
# EXCEL REPORT WRITER
# ─────────────────────────────────────────────
HEAT_MAP_COLORS = {
    "Critical": "C00000",
    "High":     "FF0000",
    "Medium":   "FFC000",
    "Low":      "92D050",
}

def write_heat_map_sheet(ws, df, heat_map):
    """Write a visual 5x5 heat map to a worksheet."""
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin        = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "IT RISK HEAT MAP — Likelihood vs. Impact"
    ws["A1"].font      = Font(bold=True, size=14, color="1F3864")
    ws["A1"].alignment = center

    # Column headers (Impact 1-5)
    ws["A3"] = "Likelihood \\ Impact"
    ws["A3"].fill      = header_fill
    ws["A3"].font      = header_font
    ws["A3"].alignment = center
    ws["A3"].border    = thin
    ws.column_dimensions["A"].width = 22

    impact_labels_short = {1: "1-Negligible", 2: "2-Minor", 3: "3-Moderate", 4: "4-Major", 5: "5-Critical"}
    for col_idx, i in enumerate(range(1, 6), start=2):
        cell = ws.cell(row=3, column=col_idx)
        cell.value     = impact_labels_short[i]
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # Row headers (Likelihood 5 down to 1) and cells
    likelihood_labels_short = {
        5: "5-Almost Certain",
        4: "4-Likely",
        3: "3-Possible",
        2: "2-Unlikely",
        1: "1-Rare"
    }

    for row_offset, l in enumerate(range(5, 0, -1)):
        row_num = 4 + row_offset
        ws.row_dimensions[row_num].height = 50

        # Row header
        hdr = ws.cell(row=row_num, column=1)
        hdr.value     = likelihood_labels_short[l]
        hdr.fill      = header_fill
        hdr.font      = header_font
        hdr.alignment = center
        hdr.border    = thin

        for col_idx, i in enumerate(range(1, 6), start=2):
            score        = l * i
            rating, color = get_risk_rating(score)
            risks_here   = heat_map.get((l, i), [])

            cell         = ws.cell(row=row_num, column=col_idx)
            cell.value   = f"Score: {score}\n" + ("\n".join(risks_here) if risks_here else "")
            cell.fill    = PatternFill("solid", fgColor=color)
            cell.font    = Font(bold=True, color="FFFFFF" if rating in ["Critical", "High"] else "000000", size=9)
            cell.alignment = center
            cell.border  = thin

    # Legend
    legend_row = 10
    ws.cell(row=legend_row, column=1).value = "LEGEND"
    ws.cell(row=legend_row, column=1).font  = Font(bold=True, size=11)

    for idx, (low, high, label, color) in enumerate(RISK_RATING_BANDS):
        r = legend_row + 1 + idx
        ws.cell(row=r, column=1).value     = f"{label} ({low}–{high})"
        ws.cell(row=r, column=1).fill      = PatternFill("solid", fgColor=color)
        ws.cell(row=r, column=1).font      = Font(
            bold=True, color="FFFFFF" if label in ["Critical", "High"] else "000000"
        )
        ws.cell(row=r, column=1).alignment = center
        ws.cell(row=r, column=1).border    = thin


def style_register_sheet(ws):
    """Apply formatting to the scored risk register sheet."""
    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    col_widths = [6, 10, 45, 20, 12, 8, 20, 8, 12, 20, 22, 14, 14, 14, 14, 14, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for cell in ws[1]:
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border
    ws.row_dimensions[1].height = 30

    # Color-code residual rating column
    residual_rating_col = None
    inherent_rating_col = None
    for i, cell in enumerate(ws[1], 1):
        if cell.value == "residual_rating":
            residual_rating_col = i
        if cell.value == "inherent_rating":
            inherent_rating_col = i

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border    = thin_border
            cell.alignment = left_align

        for col_idx in [residual_rating_col, inherent_rating_col]:
            if col_idx:
                rc = row[col_idx - 1]
                rating_color = {
                    "Critical": ("C00000", "FFFFFF"),
                    "High":     ("FF0000", "FFFFFF"),
                    "Medium":   ("FFC000", "000000"),
                    "Low":      ("92D050", "000000"),
                }.get(str(rc.value), None)
                if rating_color:
                    rc.fill      = PatternFill("solid", fgColor=rating_color[0])
                    rc.font      = Font(bold=True, color=rating_color[1])
                    rc.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def write_report(df_scored, df_exceptions, heat_map, output_path, fmt="xlsx"):
    """Write full report — Excel or CSV."""
    today = datetime.today()

    # Executive summary stats
    total       = len(df_scored)
    critical    = len(df_scored[df_scored["residual_rating"] == "Critical"])
    high        = len(df_scored[df_scored["residual_rating"] == "High"])
    medium      = len(df_scored[df_scored["residual_rating"] == "Medium"])
    low         = len(df_scored[df_scored["residual_rating"] == "Low"])
    no_controls = len(df_scored[df_scored["control_effectiveness"].apply(normalize_str).isin(["none", ""])])
    no_owner    = len(df_scored[df_scored["owner"].apply(normalize_str) == ""])
    overdue     = len(df_exceptions[df_exceptions["Exception Type"].str.contains("Overdue", na=False)]) \
                  if not df_exceptions.empty else 0

    summary_rows = [
        ("Report Date",              today.strftime("%Y-%m-%d")),
        ("Total Risks",              total),
        ("Critical (Score 20–25)",   critical),
        ("High (Score 10–19)",       high),
        ("Medium (Score 5–9)",       medium),
        ("Low (Score 1–4)",          low),
        ("Risks with No Controls",   no_controls),
        ("Risks with No Owner",      no_owner),
        ("Overdue Mitigations",      overdue),
        ("Total Exceptions Flagged", len(df_exceptions) if not df_exceptions.empty else 0),
        ("Avg Inherent Score",       round(df_scored["inherent_score"].mean(), 1)),
        ("Avg Residual Score",       round(df_scored["residual_score"].mean(), 1)),
        ("Avg Risk Reduction",       f"{round(df_scored['risk_reduction_%'].mean(), 0):.0f}%"),
        ("Prepared By",              ""),
        ("Reviewed By",              ""),
    ]
    df_summary = pd.DataFrame(summary_rows, columns=["Metric", "Value"])

    # Prepare display copy
    display_cols = [
        "rank", "risk_id", "risk_description", "category",
        "likelihood", "likelihood_label", "impact", "impact_label",
        "inherent_score", "inherent_rating",
        "control_description", "control_effectiveness", "control_factor", "risk_reduction_%",
        "residual_score", "residual_rating",
        "owner", "mitigation_due_date", "status"
    ]
    df_display = df_scored[[c for c in display_cols if c in df_scored.columns]].copy()
    if "mitigation_due_date" in df_display.columns:
        df_display["mitigation_due_date"] = pd.to_datetime(
            df_display["mitigation_due_date"], errors="coerce"
        ).dt.strftime("%Y-%m-%d").fillna("")

    empty_exc_cols = ["Risk ID", "Description", "Category", "Residual Score",
                      "Residual Rating", "Exception Type", "Severity", "Detail", "Recommendation"]

    if fmt == "csv":
        base = output_path.replace(".csv", "").replace(".xlsx", "")
        paths = {
            "scored_register": f"{base}_scored_register.csv",
            "exceptions":      f"{base}_exceptions.csv",
            "summary":         f"{base}_summary.csv",
        }
        df_display.to_csv(paths["scored_register"], index=False)
        exc_df = df_exceptions if not df_exceptions.empty else pd.DataFrame(columns=empty_exc_cols)
        exc_df.to_csv(paths["exceptions"], index=False)
        df_summary.to_csv(paths["summary"], index=False)
        print("   CSV files written:")
        for _, path in paths.items():
            print(f"   ✓ {path}")
    else:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_display.to_excel(writer, sheet_name="Risk Register", index=False)
            exc_df = df_exceptions if not df_exceptions.empty else pd.DataFrame(columns=empty_exc_cols)
            exc_df.to_excel(writer, sheet_name="Exceptions", index=False)
            df_summary.to_excel(writer, sheet_name="Executive Summary", index=False)
            # Add blank heat map sheet (styled after)
            pd.DataFrame().to_excel(writer, sheet_name="Heat Map", index=False)

        # Post-process: style sheets and draw heat map
        wb = load_workbook(output_path)

        # Style Risk Register
        style_register_sheet(wb["Risk Register"])

        # Draw heat map
        write_heat_map_sheet(wb["Heat Map"], df_scored, heat_map)

        # Style Exceptions
        ws_exc = wb["Exceptions"]
        hf     = PatternFill("solid", fgColor="1F3864")
        hfont  = Font(bold=True, color="FFFFFF", size=11)
        thin   = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        for cell in ws_exc[1]:
            cell.fill = hf; cell.font = hfont; cell.alignment = center; cell.border = thin
        exc_widths = [10, 40, 16, 14, 16, 32, 12, 45, 50]
        for i, w in enumerate(exc_widths, 1):
            ws_exc.column_dimensions[get_column_letter(i)].width = w

        # Color severity in exceptions
        sev_col = next((i for i, c in enumerate(ws_exc[1], 1) if c.value == "Severity"), None)
        for row in ws_exc.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin; cell.alignment = left
            if sev_col:
                sc = row[sev_col - 1]
                color_map = {"Critical": ("C00000","FFFFFF"), "High": ("FF0000","FFFFFF"),
                             "Medium": ("FFC000","000000"), "Low": ("92D050","000000")}
                cm = color_map.get(str(sc.value))
                if cm:
                    sc.fill = PatternFill("solid", fgColor=cm[0])
                    sc.font = Font(bold=True, color=cm[1])
                    sc.alignment = center

        if ws_exc.dimensions != "A1:A1":
            ws_exc.freeze_panes = "A2"
            ws_exc.auto_filter.ref = ws_exc.dimensions

        # Style Executive Summary
        ws_sum = wb["Executive Summary"]
        for cell in ws_sum[1]:
            cell.fill = hf; cell.font = hfont; cell.alignment = center; cell.border = thin
        for row in ws_sum.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin; cell.alignment = left
        ws_sum.column_dimensions["A"].width = 30
        ws_sum.column_dimensions["B"].width = 20

        # Reorder sheets
        desired_order = ["Risk Register", "Heat Map", "Exceptions", "Executive Summary"]
        for i, name in enumerate(desired_order):
            if name in wb.sheetnames:
                wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

        wb.save(output_path)

    return df_summary


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Risk Score Calculator — IT Audit / GRC Tool")
    parser.add_argument("--input",  "-i", default=None,
                        help="Path to risk register CSV")
    parser.add_argument("--output", "-o", default="risk_score_report.xlsx",
                        help="Output file path (default: risk_score_report.xlsx)")
    parser.add_argument("--format", "-f", choices=["xlsx", "csv"], default="xlsx",
                        help="Output format: xlsx (default) or csv")
    args = parser.parse_args()

    print("=" * 60)
    print("  RISK SCORE CALCULATOR — GRC & IT Audit Toolkit")
    print("=" * 60)

    # Load data
    if args.input:
        print(f"\n📂 Loading: {args.input}")
        df = pd.read_csv(args.input)
        try:
            validate_input(df)
        except ValueError as e:
            print(f"\n❌ ERROR: {e}")
            return
    else:
        print("\n⚠️  No input file provided. Using sample data for demo.")
        df = generate_sample_data()

    print(f"   {len(df)} risks loaded.")

    # Score risks
    print("\n📐 Calculating risk scores...")
    print("   ✓ Inherent score  = Likelihood × Impact")
    print("   ✓ Residual score  = Inherent × (1 − Control Effectiveness)")
    print("   ✓ Risk reduction  = Control Effectiveness %")
    df_scored = calculate_scores(df)

    # Detect exceptions
    print("\n🔍 Scanning for risk governance exceptions...")
    checks = [
        "No owner assigned",
        "No controls in place",
        "Overdue mitigations",
        "High/Critical residual risk with weak controls",
        "Stale open risks (90+ days overdue)",
    ]
    for c in checks:
        print(f"   ✓ {c}")
    df_exceptions = detect_exceptions(df_scored)

    # Build heat map data
    heat_map = build_heat_map_data(df_scored)

    # Write report
    print(f"\n📊 Writing report: {args.output}")
    write_report(df_scored, df_exceptions, heat_map, args.output, fmt=args.format)

    # Console summary
    print("\n" + "=" * 60)
    print("  RISK SCORE SUMMARY")
    print("=" * 60)
    for rating in ["Critical", "High", "Medium", "Low"]:
        count = len(df_scored[df_scored["residual_rating"] == rating])
        if count:
            print(f"   {rating:<12}: {count} risk(s)")

    print(f"\n   Avg Inherent Score : {df_scored['inherent_score'].mean():.1f}")
    print(f"   Avg Residual Score : {df_scored['residual_score'].mean():.1f}")
    print(f"   Avg Risk Reduction : {df_scored['risk_reduction_%'].mean():.0f}%")

    if not df_exceptions.empty:
        print(f"\n   Exceptions Found   : {len(df_exceptions)}")
        for sev in ["Critical", "High", "Medium"]:
            count = len(df_exceptions[df_exceptions["Severity"] == sev])
            if count:
                print(f"   {sev:<12}   : {count}")

    if args.format == "csv":
        base = args.output.replace(".csv","").replace(".xlsx","")
        print(f"\n✅ CSV files saved: {base}_*.csv")
    else:
        print(f"\n✅ Report saved: {args.output}")
        print("   Sheets: Risk Register | Heat Map | Exceptions | Executive Summary")
    print("=" * 60)


if __name__ == "__main__":
    main()
