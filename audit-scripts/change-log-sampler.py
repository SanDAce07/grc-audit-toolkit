"""
Change Log Sampler
==================
IT Audit Tool — GRC Toolkit
Author: Sandesh Lama Tamang | ULM Accounting & CIS

Purpose:
    Ingests a change management log CSV, applies an illustrative risk-based
    portfolio sampling rule, pre-flags exceptions, and outputs
    a formatted Excel workpaper ready for audit testing.

Illustrative default logic (not an auditing-standard sample-size table):
    Population 1–24    → Test all
    Population 25–99   → Sample 25
    Population 100+    → Sample 40
    Emergency changes  → Include all as targeted selections for this scenario

    Professional engagements should determine sample size from the governing
    methodology, control frequency, tolerable and expected deviation, sampling
    risk, population characteristics, and auditor judgment. Use --sample-size
    to supply that documented engagement-specific decision.

Input CSV columns (required):
    change_id, description, category, requested_by, approved_by,
    implemented_by, request_date, approval_date, implementation_date,
    status, is_emergency, tested

Usage:
    python change-log-sampler.py --input change_log.csv --output change_log_sample.xlsx
    python change-log-sampler.py  (uses sample data if no file provided)
    python change-log-sampler.py --sample-size 25 --seed 42
"""

import pandas as pd
import argparse
import random
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

REQUIRED_COLUMNS = [
    "change_id", "description", "category", "requested_by", "approved_by",
    "implemented_by", "request_date", "approval_date", "implementation_date",
    "status", "is_emergency", "tested"
]

# ─────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────
def normalize_str(value):
    """Normalize string for comparison — null-safe, stripped, lowercase."""
    if value is None:
        return ""
    if pd.isna(value) if not isinstance(value, str) else False:
        return ""
    return str(value).strip().lower()


# Illustrative portfolio defaults only; not an AICPA/PCAOB sample-size table.
SAMPLE_THRESHOLDS = [
    (0,   24,  None),   # Test all
    (25,  99,  25),     # Sample 25
    (100, 999, 40),     # Sample 40
]


# ─────────────────────────────────────────────
# INPUT VALIDATION
# ─────────────────────────────────────────────
def validate_input(df):
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"Input CSV is missing required column(s): {', '.join(missing)}\n"
            f"Required columns: {', '.join(REQUIRED_COLUMNS)}"
        )
    for col in ["change_id", "requested_by", "approved_by", "implemented_by", "status"]:
        null_count = df[col].isna().sum()
        if null_count > 0:
            print(f"   ⚠️  Warning: {null_count} null value(s) in '{col}' — these rows may produce exceptions.")
    return True


# ─────────────────────────────────────────────
# SAMPLE DATA GENERATOR
# ─────────────────────────────────────────────
def generate_sample_data():
    """Generate a realistic change log with seeded anomalies."""
    today = datetime.today()
    categories = ["Normal", "Standard", "Emergency", "Routine"]
    systems    = ["ERP", "Active Directory", "Financial System", "Payroll", "Network", "Database"]
    users      = ["john.smith", "sara.jones", "mike.chen", "lisa.patel", "tom.brown",
                  "anna.kim",  "raj.patel",  "emily.wu",  "dan.garcia", "kate.moore"]

    random.seed(99)
    data = []

    for i in range(1, 76):  # 75 changes → triggers 40-sample rule
        req_date  = today - timedelta(days=random.randint(10, 365))
        app_date  = req_date  + timedelta(days=random.randint(0, 3))
        impl_date = app_date  + timedelta(days=random.randint(1, 7))

        requester   = random.choice(users)
        approver    = random.choice([u for u in users if u != requester])
        implementer = random.choice([u for u in users if u != requester])

        is_emergency = "Yes" if i % 15 == 0 else "No"
        category     = "Emergency" if is_emergency == "Yes" else random.choice(["Normal", "Standard", "Routine"])

        data.append({
            "change_id":             f"CHG-{i:04d}",
            "description":           f"Update to {random.choice(systems)} — {category.lower()} change #{i}",
            "category":              category,
            "requested_by":          requester,
            "approved_by":           approver,
            "implemented_by":        implementer,
            "request_date":          req_date.strftime("%Y-%m-%d"),
            "approval_date":         app_date.strftime("%Y-%m-%d"),
            "implementation_date":   impl_date.strftime("%Y-%m-%d"),
            "status":                "Completed",
            "is_emergency":          is_emergency,
            "tested":                random.choice(["Yes", "Yes", "Yes", "No"]),
        })

    # ── Inject specific anomalies ──

    # Missing approval
    data[4]["approved_by"]  = ""
    data[4]["approval_date"] = ""

    # SOD violation: requester = implementer
    data[9]["implemented_by"] = data[9]["requested_by"]

    # Same-day approval and implementation (rushed change)
    data[14]["approval_date"]      = data[14]["request_date"]
    data[14]["implementation_date"] = data[14]["request_date"]

    # Implemented before approval
    base = today - timedelta(days=50)
    data[19]["approval_date"]       = (base + timedelta(days=5)).strftime("%Y-%m-%d")
    data[19]["implementation_date"] = (base + timedelta(days=2)).strftime("%Y-%m-%d")

    # Not tested
    data[24]["tested"] = "No"
    data[29]["tested"] = "No"

    # SOD: approver = implementer
    data[34]["approved_by"] = data[34]["implemented_by"]

    # Missing implementation date
    data[39]["implementation_date"] = ""
    data[39]["status"]              = "Open"

    return pd.DataFrame(data)


# ─────────────────────────────────────────────
# SAMPLING ENGINE
# ─────────────────────────────────────────────
def determine_sample_size(population, override=None):
    """Return the explicit override or the documented portfolio-demo default."""
    if override:
        actual = min(override, population)
        if actual < override:
            print(f"   \u26a0\ufe0f  Sample size override ({override}) exceeds population ({population}). "
                  f"Using {actual} instead.")
        return actual
    for low, high, size in SAMPLE_THRESHOLDS:
        if low <= population <= high:
            return population if size is None else size
    return 40  # default for very large populations


def select_sample(df, sample_size, seed):
    """
    Select audit sample:
    - All emergency changes (targeted selections for this scenario)
    - Random selection from remainder to hit sample_size
    Returns (sample_df, method_description)
    """
    if df.empty:
        raise ValueError("Cannot sample from empty dataset.")
    if sample_size <= 0:
        raise ValueError("Sample size must be greater than 0.")
    emergency = df[df["is_emergency"].str.strip().str.lower() == "yes"].copy()
    normal    = df[df["is_emergency"].str.strip().str.lower() != "yes"].copy()

    emergency["selection_method"] = "Targeted — Emergency Change"
    selected_normal = pd.DataFrame()

    remaining_needed = max(0, sample_size - len(emergency))

    if remaining_needed > 0 and len(normal) > 0:
        random.seed(seed)
        n = min(remaining_needed, len(normal))
        selected_normal = normal.sample(n=n, random_state=seed).copy()
        selected_normal["selection_method"] = f"Random Sample (seed={seed})"

    sample = pd.concat([emergency, selected_normal]).sort_values("change_id").reset_index(drop=True)
    sample["sample_number"] = range(1, len(sample) + 1)

    method = (
        f"Illustrative risk-based sampling | "
        f"Population: {len(df)} | "
        f"Sample Size: {sample_size} | "
        f"Emergency (targeted): {len(emergency)} | "
        f"Random: {len(selected_normal)} | "
        f"Random seed: {seed}"
    )
    return sample, method


# ─────────────────────────────────────────────
# EXCEPTION DETECTION ENGINE
# ─────────────────────────────────────────────
def detect_exceptions(df):
    """Scan all changes (not just sample) for automatic red flags."""
    exceptions = []

    date_cols = ["request_date", "approval_date", "implementation_date"]
    for col in date_cols:
        df[col] = pd.to_datetime(df[col], errors="coerce")
        invalid = df[df[col].isna() & df[col.replace("_date","_date")].notna()] if False else \
                  df[pd.to_datetime(df[col], errors="coerce").isna() & df[col].notna()]
        # Re-check original raw values for invalid (non-empty but unparseable) dates
        raw_invalid = df[df[col].isna()].shape[0]
        if raw_invalid > 0:
            print(f"   \u26a0\ufe0f  Warning: {raw_invalid} unparseable/missing value(s) in \'{col}\' "
                  f"\u2014 these rows will skip date-based checks.")

    for _, row in df.iterrows():
        cid = row["change_id"]

        def flag(exc_type, risk, detail, rec):
            exceptions.append({
                "Change ID":      cid,
                "Description":    row["description"],
                "Category":       row["category"],
                "Exception Type": exc_type,
                "Risk Rating":    risk,
                "Detail":         detail,
                "Recommendation": rec,
            })

        # 1. Missing approval — independent if (not elif)
        if normalize_str(row["approved_by"]) == "":
            flag(
                "Missing Approval",
                "Critical",
                f"Change {cid} has no recorded approver.",
                "Obtain approval documentation or flag as unauthorized change. Escalate to change manager."
            )

        # 2. SOD — requester is implementer (independent check, catches overlap with #1)
        if normalize_str(row["requested_by"]) == normalize_str(row["implemented_by"]) \
                and normalize_str(row["requested_by"]) != "":
            flag(
                "SOD Violation — Requester = Implementer",
                "High",
                f"{row['requested_by']} both requested and implemented change {cid}.",
                "Require independent implementation. Update SDLC policy to enforce segregation."
            )

        # 3. SOD — approver is implementer (independent check)
        if normalize_str(row["approved_by"]) == normalize_str(row["implemented_by"]) \
                and normalize_str(row["approved_by"]) != "":
            flag(
                "SOD Violation — Approver = Implementer",
                "High",
                f"{row['approved_by']} both approved and implemented change {cid}.",
                "Require independent approval from someone who did not implement the change."
            )

        # 4. Implemented before approval
        if pd.notna(row["approval_date"]) and pd.notna(row["implementation_date"]):
            if row["implementation_date"] < row["approval_date"]:
                flag(
                    "Implemented Before Approval",
                    "Critical",
                    f"Change {cid} implemented on {row['implementation_date'].date()} "
                    f"before approval on {row['approval_date'].date()}.",
                    "Investigate root cause. Determine if change was authorized retroactively. "
                    "Strengthen pre-implementation approval gate controls."
                )

        # 5. Same-day request, approval, and implementation (rushed)
        if pd.notna(row["request_date"]) and pd.notna(row["approval_date"]) \
                and pd.notna(row["implementation_date"]):
            if row["request_date"] == row["approval_date"] == row["implementation_date"] \
                    and str(row["is_emergency"]).strip().lower() != "yes":
                flag(
                    "Same-Day Change — Not Emergency",
                    "Medium",
                    f"Change {cid} was requested, approved, and implemented on the same day "
                    f"({row['request_date'].date()}) but is not flagged as an emergency.",
                    "Verify adequate review time was provided. Consider if this should be reclassified "
                    "as an emergency change or if the approval process was circumvented."
                )

        # 6. Not tested
        if str(row["tested"]).strip().lower() in ["no", "false", "0", ""]:
            flag(
                "Change Not Tested",
                "High",
                f"Change {cid} has no recorded testing evidence.",
                "Obtain testing documentation (UAT sign-off, test results). "
                "If not tested, assess risk and determine if rollback is required."
            )

        # 7. Missing implementation date on completed change
        if str(row["status"]).strip().lower() == "completed" \
                and pd.isna(row["implementation_date"]):
            flag(
                "Completed Change — No Implementation Date",
                "Medium",
                f"Change {cid} is marked Completed but has no implementation date recorded.",
                "Update change record with implementation date. Review change management "
                "process to ensure all fields are completed before closure."
            )

    return pd.DataFrame(exceptions) if exceptions else pd.DataFrame()


# ─────────────────────────────────────────────
# ADD WORKPAPER COLUMNS TO SAMPLE
# ─────────────────────────────────────────────
def add_workpaper_columns(sample_df, df_exceptions):
    """Add audit testing columns without treating unrelated flags as missing approval."""
    wp = sample_df.copy()

    exception_ids = set(df_exceptions["Change ID"].tolist()) if not df_exceptions.empty else set()
    missing_approval_ids = set(
        df_exceptions.loc[
            df_exceptions["Exception Type"] == "Missing Approval", "Change ID"
        ].tolist()
    ) if not df_exceptions.empty else set()

    # Pre-populate exception column based on auto-detection
    def get_exception_note(cid):
        if cid in exception_ids:
            return "⚠️ See Exceptions Sheet"
        return ""

    wp["Approval Documented?"]    = wp.apply(
        lambda row: "No ⚠️" if row["change_id"] in missing_approval_ids
        else ("Yes" if normalize_str(row["approved_by"]) else ""),
        axis=1,
    )
    wp["Testing Evidence?"]       = ""
    wp["SOD Verified?"]           = ""
    wp["Impl. Date Matches?"]     = ""
    wp["Auditor Notes"]           = wp["change_id"].apply(get_exception_note)
    wp["Workpaper Conclusion"]    = ""   # Effective / Exception
    wp["Auditor Initials"]        = ""
    wp["Review Date"]             = ""

    return wp


# ─────────────────────────────────────────────
# EXCEL REPORT WRITER
# ─────────────────────────────────────────────
RISK_COLORS = {
    "Critical": "C00000",
    "High":     "FF0000",
    "Medium":   "FFC000",
    "Low":      "92D050",
}

def style_workbook(path, has_exceptions):
    wb = load_workbook(path)

    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    def format_sheet(ws, col_widths):
        for cell in ws[1]:
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = center_align
            cell.border    = thin_border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border    = thin_border
                cell.alignment = left_align
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 30

    sheet_configs = {
        "Full Change Log": [12, 40, 14, 16, 16, 16, 14, 14, 18, 12, 14, 10],
        "Selected Sample": [8, 12, 40, 14, 14, 22, 20, 18, 18, 18, 35, 20, 16, 14],
        "Sample Evidence": [8, 12, 16, 16, 16, 14, 14, 18, 12, 10],
        "Exceptions":      [12, 40, 14, 32, 12, 45, 50],
        "Summary":         [40, 20],
    }

    for sheet_name, widths in sheet_configs.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        format_sheet(ws, widths)
        ws.freeze_panes = "A2"
        if ws.dimensions != "A1:A1":
            ws.auto_filter.ref = ws.dimensions

    if "Selected Sample" in wb.sheetnames:
        ws_samp = wb["Selected Sample"]
        ws_samp.freeze_panes = "C2"
        ws_samp.sheet_view.zoomScale = 85
        ws_samp.sheet_properties.pageSetUpPr.fitToPage = True
        ws_samp.page_setup.orientation = "landscape"
        ws_samp.page_setup.fitToWidth = 1
        ws_samp.page_setup.fitToHeight = 0
        ws_samp.print_title_rows = "1:1"

    # Color-code risk ratings in Exceptions sheet
    if "Exceptions" in wb.sheetnames and has_exceptions:
        ws_exc = wb["Exceptions"]
        risk_col = next(
            (i for i, cell in enumerate(ws_exc[1], 1) if cell.value == "Risk Rating"), None
        )
        if risk_col:
            for row in ws_exc.iter_rows(min_row=2):
                cell = row[risk_col - 1]
                color = RISK_COLORS.get(str(cell.value))
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(
                        bold=True,
                        color="FFFFFF" if cell.value in ["Critical", "High"] else "000000"
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Highlight workpaper conclusion column in sample sheet
    if "Selected Sample" in wb.sheetnames:
        ws_samp = wb["Selected Sample"]
        conc_col = next(
            (i for i, c in enumerate(ws_samp[1], 1) if c.value == "Workpaper Conclusion"), None
        )
        if conc_col:
            wp_fill = PatternFill("solid", fgColor="FFF2CC")  # Light yellow — needs completion
            for row in ws_samp.iter_rows(min_row=2):
                cell = row[conc_col - 1]
                if not cell.value:
                    cell.fill = wp_fill

    wb.save(path)


def write_report(df_all, df_sample, df_exceptions, method, output_path, fmt="xlsx"):
    """Write all sheets to formatted Excel workbook or CSV files."""

    # Prepare date columns for display
    date_cols = ["request_date", "approval_date", "implementation_date"]
    def fmt_dates(df):
        d = df.copy()
        for col in date_cols:
            if col in d.columns:
                d[col] = pd.to_datetime(d[col], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")
        return d

    df_all_display    = fmt_dates(df_all)
    df_sample_display = fmt_dates(df_sample)

    workpaper_columns = [
        "sample_number", "change_id", "description", "category", "is_emergency",
        "selection_method", "Approval Documented?", "Testing Evidence?",
        "SOD Verified?", "Impl. Date Matches?", "Auditor Notes",
        "Workpaper Conclusion", "Auditor Initials", "Review Date",
    ]
    evidence_columns = [
        "sample_number", "change_id", "requested_by", "approved_by", "implemented_by",
        "request_date", "approval_date", "implementation_date", "status", "tested",
    ]
    df_sample_workpaper = df_sample_display[workpaper_columns]
    df_sample_evidence = df_sample_display[evidence_columns]

    # Summary data
    summary_rows = [
        ("Audit Period",              "Full CSV range"),
        ("Total Population",          len(df_all)),
        ("Emergency Changes",         len(df_all[df_all["is_emergency"].str.strip().str.lower() == "yes"])),
        ("Normal Changes",            len(df_all[df_all["is_emergency"].str.strip().str.lower() != "yes"])),
        ("Sample Size Selected",      len(df_sample)),
        ("Emergency (targeted)",      len(df_sample[df_sample["selection_method"].str.contains("Emergency", na=False)])),
        ("Random Sample",             len(df_sample[df_sample["selection_method"].str.contains("Random", na=False)])),
        ("Pre-flagged Exceptions",    len(df_exceptions) if not df_exceptions.empty else 0),
        ("Sampling Method",           method),
        ("Script Run Date",           datetime.today().strftime("%Y-%m-%d")),
        ("Prepared By",               ""),
        ("Reviewed By",               ""),
    ]
    df_summary = pd.DataFrame(summary_rows, columns=["Item", "Value"])

    empty_exc_cols = ["Change ID", "Description", "Category",
                      "Exception Type", "Risk Rating", "Detail", "Recommendation"]

    if fmt == "csv":
        # Export as separate CSV files, one per sheet
        base = output_path.replace(".csv", "").replace(".xlsx", "")
        paths = {
            "full_change_log": f"{base}_full_change_log.csv",
            "selected_sample": f"{base}_selected_sample.csv",
            "sample_evidence": f"{base}_sample_evidence.csv",
            "exceptions":      f"{base}_exceptions.csv",
            "summary":         f"{base}_summary.csv",
        }
        df_all_display.to_csv(paths["full_change_log"], index=False)
        df_sample_workpaper.to_csv(paths["selected_sample"], index=False)
        df_sample_evidence.to_csv(paths["sample_evidence"], index=False)
        exc_df = df_exceptions if not df_exceptions.empty else pd.DataFrame(columns=empty_exc_cols)
        exc_df.to_csv(paths["exceptions"], index=False)
        df_summary.to_csv(paths["summary"], index=False)
        print(f"   CSV files written:")
        for label, path in paths.items():
            print(f"   ✓ {path}")
    else:
        # Default: Excel workbook
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_all_display.to_excel(writer,    sheet_name="Full Change Log", index=False)
            df_sample_workpaper.to_excel(writer, sheet_name="Selected Sample", index=False)
            df_sample_evidence.to_excel(writer, sheet_name="Sample Evidence", index=False)
            exc_df = df_exceptions if not df_exceptions.empty else pd.DataFrame(columns=empty_exc_cols)
            exc_df.to_excel(writer, sheet_name="Exceptions", index=False)
            df_summary.to_excel(writer, sheet_name="Summary", index=False)
        style_workbook(output_path, not df_exceptions.empty)

    return df_summary


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Change Log Sampler — IT Audit Tool")
    parser.add_argument("--input",       "-i", default=None,
                        help="Path to change log CSV file")
    parser.add_argument("--output",      "-o", default="change_log_sample.xlsx",
                        help="Output Excel file path")
    parser.add_argument("--sample-size", "-n", type=int, default=None,
                        help="Engagement-specific sample size override (default: illustrative portfolio rule)")
    parser.add_argument("--seed",        "-s", type=int, default=42,
                        help="Random seed for reproducible sampling (default: 42)")
    parser.add_argument("--format",      "-f", choices=["xlsx", "csv"], default="xlsx",
                        help="Output format: xlsx (default) or csv (exports 5 separate CSV files)")
    args = parser.parse_args()

    print("=" * 60)
    print("  CHANGE LOG SAMPLER — GRC & IT Audit Toolkit")
    print("=" * 60)

    # Load data
    if args.input:
        print(f"\n📂 Loading: {args.input}")
        df = pd.read_csv(args.input)
        try:
            validate_input(df)
        except ValueError as e:
            print(f"\n❌ ERROR: {e}")
            return
    else:
        print("\n⚠️  No input file provided. Using sample data for demo.")
        df = generate_sample_data()

    print(f"   {len(df)} change records loaded.")

    # Determine sample size
    sample_size = determine_sample_size(len(df), args.sample_size)
    print(f"\n📐 Sampling Strategy (illustrative portfolio rule):")
    print(f"   Population   : {len(df)}")
    print(f"   Sample Size  : {sample_size}")
    print(f"   Random Seed  : {args.seed}")

    # Select sample
    df_sample, method = select_sample(df, sample_size, args.seed)
    print(f"   Emergency    : {len(df_sample[df_sample['selection_method'].str.contains('Emergency', na=False)])} (targeted 100% for this scenario)")
    print(f"   Random       : {len(df_sample[df_sample['selection_method'].str.contains('Random', na=False)])}")

    # Detect exceptions across full population
    print(f"\n🔍 Scanning full population for exceptions...")
    df_exceptions = detect_exceptions(df)

    checks = [
        "Missing approval",
        "SOD violations (requester = implementer)",
        "SOD violations (approver = implementer)",
        "Implemented before approval",
        "Same-day non-emergency changes",
        "Changes not tested",
        "Completed changes missing implementation date",
    ]
    for c in checks:
        print(f"   ✓ {c}")

    # Add workpaper columns
    df_sample = add_workpaper_columns(df_sample, df_exceptions)

    # Write report
    print(f"\n📊 Writing report: {args.output}")
    write_report(df, df_sample, df_exceptions, method, args.output, fmt=args.format)

    # Console summary
    print("\n" + "=" * 60)
    print("  RESULTS SUMMARY")
    print("=" * 60)
    print(f"   Population       : {len(df)}")
    print(f"   Sample Selected  : {len(df_sample)}")
    if not df_exceptions.empty:
        print(f"\n   Exceptions Found : {len(df_exceptions)}")
        for risk in ["Critical", "High", "Medium", "Low"]:
            count = len(df_exceptions[df_exceptions["Risk Rating"] == risk])
            if count > 0:
                print(f"   {risk:<12}   : {count}")
    else:
        print(f"\n   ✅ No exceptions detected in full population.")

    if args.format == "csv":
        base = args.output.replace(".csv", "").replace(".xlsx", "")
        print(f"\n✅ CSV files saved with prefix: {base}_*.csv")
        print(f"   Files: full_change_log | selected_sample | sample_evidence | exceptions | summary")
    else:
        print(f"\n✅ Report saved: {args.output}")
        print(f"   Sheets: Full Change Log | Selected Sample | Sample Evidence | Exceptions | Summary")
    print("=" * 60)


if __name__ == "__main__":
    main()
