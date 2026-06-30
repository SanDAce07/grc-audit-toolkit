"""
Aging Report Analyzer
=====================
IT Audit Tool — GRC Toolkit
Author: Sandesh Lama Tamang | ULM Accounting & CIS

Purpose:
    Ingests an accounts receivable aging CSV, classifies balances into aging
    buckets, flags receivable exceptions, and outputs a formatted Excel
    workbook for audit analytics and workpaper support.

Audit Logic:
    Current      -> Not yet due
    1-30 Days    -> Recently overdue
    31-60 Days   -> Moderate overdue exposure
    61-90 Days   -> High overdue exposure
    91-120 Days  -> Significant collection risk
    120+ Days    -> Severe aging / potential impairment

Input CSV columns (required):
    customer, invoice_number, invoice_date, due_date, balance, status

Accepted aliases include:
    customer name / client / account
    invoice / invoice # / document number
    date / document date
    due / payment due date
    amount / open amount / invoice amount

Usage:
    python aging-report-analyzer.py --input ar_aging.csv --output aging_report.xlsx
    python aging-report-analyzer.py  (uses sample data if no file provided)
    python aging-report-analyzer.py --as-of 2026-06-30 --top 15
"""

import argparse
import os
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REQUIRED_COLUMN_GROUPS = {
    "customer": ["customer", "customer name", "client", "account"],
    "invoice_number": ["invoice_number", "invoice", "invoice number", "invoice #", "document number"],
    "invoice_date": ["invoice_date", "date", "document date"],
    "due_date": ["due_date", "due date", "due", "payment due date"],
    "balance": ["balance", "amount", "open amount", "invoice amount"],
    "status": ["status", "invoice status", "payment status"],
}

STATUS_OPEN_VALUES = {"", "open", "outstanding", "unpaid", "pending", "active"}
STATUS_CLOSED_VALUES = {"paid", "closed", "settled", "void", "cancelled", "canceled", "written off", "write off"}
AGE_BUCKETS = [
    (None, 0, "Current"),
    (1, 30, "1-30"),
    (31, 60, "31-60"),
    (61, 90, "61-90"),
    (91, 120, "91-120"),
    (121, None, "120+"),
]

RISK_COLORS = {
    "Critical": "C00000",
    "High": "FF0000",
    "Medium": "FFC000",
    "Low": "92D050",
}


def timestamp():
    return datetime.now().strftime("%H:%M:%S")


def log(message, icon=None, indent=""):
    prefix = f"[{timestamp()}]"
    if icon:
        print(f"{indent}{prefix} {icon} {message}")
    else:
        print(f"{indent}{prefix} {message}")


def parse_args():
    parser = argparse.ArgumentParser(description="Aging Report Analyzer - Audit Analytics Tool")
    parser.add_argument("--input", "-i", default=None, help="Path to AR aging CSV file")
    parser.add_argument("--output", "-o", default="aging_report_analysis.xlsx", help="Output Excel file path")
    parser.add_argument("--as-of", default=datetime.today().strftime("%Y-%m-%d"), help="As-of date in YYYY-MM-DD format")
    parser.add_argument("--top", type=int, default=10, help="Top customers by overdue exposure to display")
    parser.add_argument("--format", "-f", choices=["xlsx", "csv"], default="xlsx",
                        help="Output format: xlsx (default) or csv (exports 4 separate CSV files)")
    return parser.parse_args()


def parse_date(value):
    if pd.isna(value) or str(value).strip() == "":
        return pd.NaT
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return pd.NaT


def parse_amount(value):
    if pd.isna(value):
        return 0.0
    cleaned = str(value).strip().replace(",", "").replace("$", "")
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    try:
        return float(Decimal(cleaned))
    except (InvalidOperation, ValueError):
        return 0.0


def normalize_str(value):
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def find_column(df, aliases):
    lookup = {str(col).strip().lower(): col for col in df.columns}
    for alias in aliases:
        alias_key = alias.strip().lower()
        if alias_key in lookup:
            return lookup[alias_key]
    return None


def validate_and_normalize_input(df):
    mapped = {}
    for standard_name, aliases in REQUIRED_COLUMN_GROUPS.items():
        col = find_column(df, aliases)
        if standard_name == "invoice_date":
            mapped[standard_name] = col
            continue
        if col is None:
            raise ValueError(
                f"Input CSV is missing a required column for '{standard_name}'. "
                f"Accepted names: {', '.join(aliases)}"
            )
        mapped[standard_name] = col

    normalized = pd.DataFrame()
    normalized["customer"] = df[mapped["customer"]].fillna("Unknown Customer").astype(str).str.strip()
    normalized["invoice_number"] = df[mapped["invoice_number"]].fillna("Unknown Invoice").astype(str).str.strip()
    normalized["invoice_date"] = df[mapped["invoice_date"]].apply(parse_date) if mapped["invoice_date"] else pd.NaT
    normalized["due_date"] = df[mapped["due_date"]].apply(parse_date)
    normalized["balance"] = df[mapped["balance"]].apply(parse_amount)
    normalized["status"] = df[mapped["status"]].apply(normalize_str)

    null_due_dates = normalized["due_date"].isna().sum()
    if null_due_dates > 0:
        print(
            f"   ⚠️  Warning: {null_due_dates} row(s) have missing or invalid due dates. "
            f"Aging classification unavailable for these rows."
        )

    invalid_balance = (normalized["balance"] == 0).sum()
    if invalid_balance > 0:
        print(
            f"   ⚠️  Warning: {invalid_balance} row(s) have zero or unparseable balances. "
            f"These rows will be excluded from aging analysis."
        )

    return normalized


def validate_output_path(output_path):
    try:
        output_dir = os.path.dirname(output_path) or "."
        if not os.path.exists(output_dir):
            print(f"❌ ERROR: Output directory does not exist: {output_dir}")
            return False
        if not os.access(output_dir, os.W_OK):
            print(f"❌ ERROR: Cannot write to directory: {output_dir}")
            return False
    except Exception as exc:
        print(f"❌ ERROR: Invalid output path: {exc}")
        return False
    return True


def generate_sample_data():
    today = datetime.today()
    data = [
        {"customer": "Acadia Health Partners", "invoice_number": "INV-1001", "invoice_date": today - timedelta(days=52), "due_date": today - timedelta(days=22), "balance": 18250.00, "status": "Open"},
        {"customer": "Acadia Health Partners", "invoice_number": "INV-1002", "invoice_date": today - timedelta(days=88), "due_date": today - timedelta(days=58), "balance": 22400.00, "status": "Open"},
        {"customer": "Bayou Industrial Supply", "invoice_number": "INV-1003", "invoice_date": today - timedelta(days=130), "due_date": today - timedelta(days=100), "balance": 31750.00, "status": "Open"},
        {"customer": "Crescent Retail Group", "invoice_number": "INV-1004", "invoice_date": today - timedelta(days=10), "due_date": today + timedelta(days=20), "balance": 9600.00, "status": "Open"},
        {"customer": "Delta Municipal Services", "invoice_number": "INV-1005", "invoice_date": today - timedelta(days=175), "due_date": today - timedelta(days=145), "balance": 48120.00, "status": "Open"},
        {"customer": "Evergreen Medical Labs", "invoice_number": "INV-1006", "invoice_date": today - timedelta(days=41), "due_date": today - timedelta(days=11), "balance": 14220.00, "status": "Open"},
        {"customer": "Frontier Logistics", "invoice_number": "INV-1007", "invoice_date": today - timedelta(days=61), "due_date": today - timedelta(days=31), "balance": 12500.00, "status": "Open"},
        {"customer": "Gulf Coast Manufacturing", "invoice_number": "INV-1008", "invoice_date": today - timedelta(days=205), "due_date": today - timedelta(days=175), "balance": 59200.00, "status": "Open"},
        {"customer": "Harbor Energy Co.", "invoice_number": "INV-1009", "invoice_date": today - timedelta(days=24), "due_date": today + timedelta(days=6), "balance": 8800.00, "status": "Open"},
        {"customer": "Ivory Tech Consulting", "invoice_number": "INV-1010", "invoice_date": today - timedelta(days=95), "due_date": today - timedelta(days=65), "balance": 17110.00, "status": "Open"},
        {"customer": "Jefferson Community Care", "invoice_number": "INV-1011", "invoice_date": today - timedelta(days=150), "due_date": today - timedelta(days=120), "balance": 26500.00, "status": "Open"},
        {"customer": "Keystone Distribution", "invoice_number": "INV-1012", "invoice_date": today - timedelta(days=72), "due_date": today - timedelta(days=42), "balance": -2400.00, "status": "Open"},
        {"customer": "Lafayette Public Works", "invoice_number": "INV-1013", "invoice_date": today - timedelta(days=33), "due_date": pd.NaT, "balance": 11200.00, "status": "Open"},
        {"customer": "Magnolia Property Group", "invoice_number": "INV-1014", "invoice_date": today - timedelta(days=40), "due_date": today - timedelta(days=10), "balance": 0.00, "status": "Closed"},
    ]
    return pd.DataFrame(data)


def filter_open_items(df):
    working = df.copy()
    working["status"] = working["status"].fillna("").astype(str).str.strip().str.lower()
    # Step 1: Remove explicitly closed items
    working = working[~working["status"].isin(STATUS_CLOSED_VALUES)]
    # Step 2: Keep only recognized open statuses (fixes bug — previously kept ANY non-empty status)
    working = working[working["status"].isin(STATUS_OPEN_VALUES)]
    # Step 3: Exclude zero balances
    working = working[working["balance"] != 0].copy()
    return working


def assign_bucket(days_past_due):
    if pd.isna(days_past_due):
        return "Unknown"
    for lower, upper, label in AGE_BUCKETS:
        if lower is None and days_past_due <= upper:
            return label
        if upper is None and days_past_due >= lower:
            return label
        if lower is not None and upper is not None and lower <= days_past_due <= upper:
            return label
    return "Unknown"


def analyze_aging(df, as_of_date):
    working = df.copy()
    working["days_past_due"] = working["due_date"].apply(
        lambda due: (as_of_date - due).days if pd.notna(due) else None
    )
    working["aging_bucket"] = working["days_past_due"].apply(assign_bucket)
    working["is_overdue"] = working["days_past_due"].apply(lambda days: pd.notna(days) and days > 0)

    customer_summary = (
        working.groupby("customer")
        .agg(
            invoice_count=("invoice_number", "count"),
            total_balance=("balance", "sum"),
            overdue_balance=(
                "balance",
                lambda s: s[working.loc[s.index, "is_overdue"] & (s > 0)].sum(),
            ),
            max_days_past_due=("days_past_due", "max"),
        )
        .reset_index()
    )

    bucket_pivot = (
        working.pivot_table(
            index="customer",
            columns="aging_bucket",
            values="balance",
            aggfunc="sum",
            fill_value=0,
        )
        .reset_index()
    )

    customer_summary = customer_summary.merge(bucket_pivot, on="customer", how="left")
    for bucket in ["Current", "1-30", "31-60", "61-90", "91-120", "120+", "Unknown"]:
        if bucket not in customer_summary.columns:
            customer_summary[bucket] = 0.0

    customer_summary = customer_summary.sort_values(["overdue_balance", "total_balance"], ascending=False).reset_index(drop=True)
    customer_summary["rank"] = range(1, len(customer_summary) + 1)

    return working, customer_summary


def detect_exceptions(df):
    findings = []

    for _, row in df.iterrows():
        def flag(finding_type, risk, detail, recommendation):
            findings.append({
                "Customer": row["customer"],
                "Invoice Number": row["invoice_number"],
                "Balance": row["balance"],
                "Days Past Due": row["days_past_due"],
                "Aging Bucket": row["aging_bucket"],
                "Finding Type": finding_type,
                "Risk Rating": risk,
                "Detail": detail,
                "Recommendation": recommendation,
            })

        if pd.isna(row["due_date"]):
            flag(
                "Missing Due Date",
                "High",
                f"Invoice {row['invoice_number']} has no usable due date and cannot be aged reliably.",
                "Obtain supporting billing documentation and correct the source report before relying on the aging analysis."
            )

        if row["balance"] < 0:
            flag(
                "Credit Balance / Negative AR",
                "Medium",
                f"Invoice {row['invoice_number']} has a negative open balance of ${row['balance']:,.2f}.",
                "Investigate whether the balance reflects an unapplied cash receipt, credit memo, or posting error."
            )

        if pd.notna(row["days_past_due"]) and row["days_past_due"] > 120 and row["balance"] > 0:
            flag(
                "Severely Aged Receivable",
                "Critical",
                f"Invoice {row['invoice_number']} is {int(row['days_past_due'])} days past due with balance ${row['balance']:,.2f}.",
                "Assess collectability, review follow-up evidence, and consider allowance or write-off implications."
            )
        elif pd.notna(row["days_past_due"]) and row["days_past_due"] > 90 and row["balance"] > 0:
            flag(
                "Significantly Overdue Receivable",
                "High",
                f"Invoice {row['invoice_number']} is {int(row['days_past_due'])} days past due with balance ${row['balance']:,.2f}.",
                "Review collection activity and evaluate whether escalation procedures were followed."
            )

    return pd.DataFrame(findings) if findings else pd.DataFrame(columns=[
        "Customer", "Invoice Number", "Balance", "Days Past Due", "Aging Bucket",
        "Finding Type", "Risk Rating", "Detail", "Recommendation"
    ])


def build_summary(df, customer_summary, findings, as_of_date):
    bucket_totals = df.groupby("aging_bucket")["balance"].sum().to_dict()
    total_open_balance = df["balance"].sum()
    positive_exposure = df.loc[df["balance"] > 0, "balance"].sum()
    overdue_balance = df.loc[df["is_overdue"] & (df["balance"] > 0), "balance"].sum()

    summary_rows = [
        ("As-of Date", as_of_date.strftime("%Y-%m-%d")),
        ("Open Invoices", len(df)),
        ("Customers", df["customer"].nunique()),
        ("Total Open Balance", round(total_open_balance, 2)),
        ("Positive Receivable Exposure", round(positive_exposure, 2)),
        ("Total Overdue Positive Balance", round(overdue_balance, 2)),
        ("Current", round(bucket_totals.get("Current", 0.0), 2)),
        ("1-30", round(bucket_totals.get("1-30", 0.0), 2)),
        ("31-60", round(bucket_totals.get("31-60", 0.0), 2)),
        ("61-90", round(bucket_totals.get("61-90", 0.0), 2)),
        ("91-120", round(bucket_totals.get("91-120", 0.0), 2)),
        ("120+", round(bucket_totals.get("120+", 0.0), 2)),
        ("Unknown", round(bucket_totals.get("Unknown", 0.0), 2)),
        ("High / Critical Findings", len(findings[findings["Risk Rating"].isin(["High", "Critical"])])),
        ("Top Customer by Overdue Exposure", customer_summary.iloc[0]["customer"] if not customer_summary.empty else ""),
        ("Note", "Customer Summary shows ALL customers ranked by overdue exposure"),
        ("Prepared By", ""),
        ("Reviewed By", ""),
    ]
    return pd.DataFrame(summary_rows, columns=["Metric", "Value"])


def detect_concentration_risk(df_aging, customer_summary, threshold_pct=20.0):
    """Flag overdue concentration against positive receivable exposure."""
    findings = []
    total_balance = df_aging.loc[df_aging["balance"] > 0, "balance"].sum()
    if total_balance <= 0:
        return pd.DataFrame()

    for _, row in customer_summary.iterrows():
        if row["overdue_balance"] <= 0:
            continue
        pct = (row["overdue_balance"] / total_balance) * 100
        if pct >= threshold_pct:
            findings.append({
                "Customer":       row["customer"],
                "Invoice Number": "— (Customer Level)",
                "Balance":        round(row["overdue_balance"], 2),
                "Days Past Due":  row["max_days_past_due"],
                "Aging Bucket":   "Multiple",
                "Finding Type":   "Concentration Risk — High Overdue Exposure",
                "Risk Rating":    "Critical" if pct >= 35 else "High",
                "Detail": (
                    f"{row['customer']} represents {pct:.1f}% of positive AR "
                    f"(${row['overdue_balance']:,.2f} overdue of ${total_balance:,.2f} positive exposure). "
                    f"Threshold: {threshold_pct:.0f}%."
                ),
                "Recommendation": (
                    "Assess collectability risk and customer credit exposure. "
                    "Review collection correspondence and evaluate whether additional "
                    "allowance for doubtful accounts is warranted. Consider credit limit review."
                ),
            })

    return pd.DataFrame(findings) if findings else pd.DataFrame()


def style_workbook(path):
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    sheet_widths = {
        "Invoice Aging": [24, 16, 14, 14, 14, 12, 12, 12],
        "Customer Summary": [8, 28, 12, 14, 14, 12, 12, 12, 12, 12, 12, 12],
        "Exceptions": [24, 16, 14, 14, 12, 34, 12, 48, 54],
        "Executive Summary": [28, 18],
    }

    for sheet_name, widths in sheet_widths.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = left_align
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        if ws.dimensions != "A1:A1":
            ws.auto_filter.ref = ws.dimensions

    if "Exceptions" in wb.sheetnames:
        ws_exc = wb["Exceptions"]
        risk_col = next((i for i, c in enumerate(ws_exc[1], 1) if c.value == "Risk Rating"), None)
        if risk_col:
            for row in ws_exc.iter_rows(min_row=2):
                cell = row[risk_col - 1]
                color = RISK_COLORS.get(str(cell.value))
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)
                    cell.font = Font(bold=True, color="FFFFFF" if cell.value in ["Critical", "High"] else "000000")
                    cell.alignment = center_align

    currency_columns = {
        "Invoice Aging": [5],
        "Customer Summary": [4, 5, 6, 7, 8, 9, 10, 11, 12],
        "Exceptions": [3],
    }
    for sheet_name, columns in currency_columns.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for col_idx in columns:
            for row in ws.iter_rows(min_row=2):
                row[col_idx - 1].number_format = "$#,##0.00"

    wb.save(path)


def write_report(df_aging, customer_summary, findings, summary_df, output_path, top_n, fmt="xlsx"):
    # Full customer list — not truncated to top_n (top_n noted in summary only)
    customer_output = customer_summary.copy()

    invoice_output = df_aging[[
        "customer", "invoice_number", "invoice_date", "due_date", "balance",
        "status", "days_past_due", "aging_bucket"
    ]].copy()
    for col in ["invoice_date", "due_date"]:
        invoice_output[col] = pd.to_datetime(invoice_output[col], errors="coerce").dt.strftime("%Y-%m-%d").fillna("")

    customer_output = customer_output[[
        "rank", "customer", "invoice_count", "total_balance", "overdue_balance", "Current",
        "1-30", "31-60", "61-90", "91-120", "120+", "Unknown"
    ]]

    if fmt == "csv":
        base = output_path.replace(".csv", "").replace(".xlsx", "")
        paths = {
            "invoice_aging":     f"{base}_invoice_aging.csv",
            "customer_summary":  f"{base}_customer_summary.csv",
            "exceptions":        f"{base}_exceptions.csv",
            "executive_summary": f"{base}_executive_summary.csv",
        }
        invoice_output.to_csv(paths["invoice_aging"], index=False)
        customer_output.to_csv(paths["customer_summary"], index=False)
        findings.to_csv(paths["exceptions"], index=False)
        summary_df.to_csv(paths["executive_summary"], index=False)
        print("   CSV files written:")
        for _, path in paths.items():
            print(f"   ✓ {path}")
    else:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            invoice_output.to_excel(writer, sheet_name="Invoice Aging", index=False)
            customer_output.to_excel(writer, sheet_name="Customer Summary", index=False)
            findings.to_excel(writer, sheet_name="Exceptions", index=False)
            summary_df.to_excel(writer, sheet_name="Executive Summary", index=False)
        style_workbook(output_path)


def main():
    args = parse_args()

    try:
        as_of_date = datetime.strptime(args.as_of, "%Y-%m-%d")
    except ValueError:
        print("❌ ERROR: --as-of must use YYYY-MM-DD format (e.g., 2026-06-29).")
        return

    if args.top < 1:
        print("❌ ERROR: --top must be a positive integer.")
        return

    if not validate_output_path(args.output):
        return

    print("=" * 60)
    print("  AGING REPORT ANALYZER — GRC & IT Audit Toolkit")
    print("=" * 60)

    if args.input:
        log(f"Loading: {args.input}", icon="📂")
        try:
            raw_df = pd.read_csv(args.input)
        except FileNotFoundError:
            print(f"\n❌ ERROR: File not found: {args.input}")
            print("   Check the path and try again.")
            return
        except Exception as exc:
            print(f"\n❌ ERROR: Could not read CSV: {exc}")
            return

        try:
            df = validate_and_normalize_input(raw_df)
        except ValueError as exc:
            print(f"\n❌ ERROR: {exc}")
            return
    else:
        log("No input file provided. Using sample data for demo.", icon="⚠️")
        df = generate_sample_data()

    df = filter_open_items(df)
    if df.empty:
        print("\n❌ ERROR: No open receivable items found after filtering.")
        print("   Check status values and balance amounts in your source data.")
        return
    print(f"   {len(df)} open receivable records loaded.")

    log("Running aging analysis...", icon="🔍")
    print("   ✓ Aging bucket classification")
    print("   ✓ Overdue customer concentration")
    print("   ✓ Missing due dates")
    print("   ✓ Negative balances / credits")
    print("   ✓ Severely aged receivables")
    print("   ✓ Customer concentration risk (>20% of total AR)")

    df_aging, customer_summary = analyze_aging(df, as_of_date)
    findings = detect_exceptions(df_aging)

    # Concentration risk — customer-level check across full AR population
    concentration_findings = detect_concentration_risk(df_aging, customer_summary)
    if not concentration_findings.empty:
        findings = pd.concat([findings, concentration_findings], ignore_index=True)

    summary_df = build_summary(df_aging, customer_summary, findings, as_of_date)

    if not findings.empty:
        print(f"\n🚨 Exceptions detected: {len(findings)}")
        for severity in ["Critical", "High", "Medium"]:
            count = len(findings[findings["Risk Rating"] == severity])
            if count:
                print(f"   {severity:<12}: {count}")
    else:
        print("\n✅ No receivable exceptions detected.")

    log(f"Writing report: {args.output}", icon="📊")
    try:
        write_report(df_aging, customer_summary, findings, summary_df, args.output, args.top, fmt=args.format)
    except PermissionError:
        print(f"\n❌ ERROR: Permission denied while writing: {args.output}")
        return
    except Exception as exc:
        print(f"\n❌ ERROR: Could not write Excel report: {exc}")
        return

    print("\n" + "=" * 60)
    print("  AGING ANALYSIS SUMMARY")
    print("=" * 60)
    print(f"   Open Invoices     : {len(df_aging)}")
    print(f"   Customers         : {df_aging['customer'].nunique()}")
    print(f"   Net AR Balance      : ${df_aging['balance'].sum():,.2f}")
    print(f"   Positive Exposure   : ${df_aging.loc[df_aging['balance'] > 0, 'balance'].sum():,.2f}")
    print(f"   Overdue Positive AR : ${df_aging.loc[df_aging['is_overdue'] & (df_aging['balance'] > 0), 'balance'].sum():,.2f}")
    if not findings.empty:
        print(f"   Exceptions Found  : {len(findings)}")
        for risk in ["Critical", "High", "Medium", "Low"]:
            count = len(findings[findings["Risk Rating"] == risk])
            if count:
                print(f"   {risk:<16}: {count}")
    else:
        print("   No receivable exceptions detected.")

    if args.format == "csv":
        base = args.output.replace(".csv", "").replace(".xlsx", "")
        print(f"\n✅ CSV files saved: {base}_*.csv")
        print("   Files: invoice_aging | customer_summary | exceptions | executive_summary")
    else:
        print(f"\n✅ Report saved: {args.output}")
        print("   Sheets: Invoice Aging | Customer Summary | Exceptions | Executive Summary")
    print("=" * 60)


if __name__ == "__main__":
    main()
